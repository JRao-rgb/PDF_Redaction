# -*- coding: utf-8 -*-
"""
Task list:
    - check multiple "words" at once, fault-tolerant string comparison
    - use names as a list instead of "first, last". Check against every element in the list
    - output as a .pdf (redact it by marking up the pdf)
    - loop through multiple files, generating "results output" file each time
"""
# %% import statements and definitions
import os
import numpy as np
import pymupdf
from pdf2image import convert_from_path
import pytesseract
import time
import re
import cv2
from tqdm import tqdm
from fillpdf import fillpdfs
import xlsxwriter
from openpyxl import load_workbook
import random

poppler_path = "C:\\Users\\jraos\\Downloads\\Release-24.08.0-0\\poppler-24.08.0\\Library\\bin"
pytesseract.pytesseract.tesseract_cmd = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

input_folder  = "C:\\Users\\jraos\\OneDrive - Stanford\\Documents\\Helping Yurui\\PDF_redaction\\inputs\\"
# input_folder = "C:\\Users\\jraos\OneDrive - Stanford\\Documents\\Helping Yurui\\PDF_redaction\\inputs_single\\"
# input_folder  = "C:\\Users\\jraos\\OneDrive - Stanford\\Documents\\Helping Yurui\\PDF_redaction\\inputs_problem\\"
# input_folder  = "C:\\Users\\jraos\\OneDrive - Stanford\\Documents\\Helping Yurui\\PDF_redaction\\inputs_2023\\"
output_folder = "C:\\Users\\jraos\\OneDrive - Stanford\\Documents\\Helping Yurui\\PDF_redaction\\outputs_complete_take_4\\"

english_ref_file = "C:\\Users\\jraos\\OneDrive - Stanford\\Documents\\Helping Yurui\\PDF_redaction\\The_Oxford_3000.txt"

# convenience parameters
page_begin = 0 # the page that we start at, in case you want to restrict the redaction to only a few pages
page_end = -1  # the page that we end at, in case you want to restrict the redaction to only a few pages
# leave page_end at -1 if you want it to analyze all pages (default)

# "solver" parameters:
dpi_used = 300 # resolution of OCR. The bigger the better, the bigger the slower. Recommend to keep it at 300.
enlarged_bounds = 10 # the number of extra pixels to search for when highlighting specific images
minimum_spacing_needed_before_categories = 50 # in searching for things like name, when do we consider it to end? Units of pixels.
minimum_number_of_redactions_before_lifting_suspicion = 2 # if a page has not enough redactions, consider it to be suspicious.
spelling_mistake_tolerance = 1 # number of difference in characters allowed for misspelled names
maximum_length_difference_tolerance = 1 # maximum difference between length of two strings before they are considered "different"
# basically, how I'm detecting names is if the name is contained in a string, we say that's name. But clearly this has its problems
# for example someone whose name is "Siva", then with spelling fault tolerant comparisons, any word that ends in -sive (e.g. hypertensive)
# will get flagged as a name. This says the maximum difiference in number of characters is 1 characters.
minimum_name_length_for_spelling_mistake_tolerant_search = 5
# the minimum length of a word needed for the spelling mistake-tolerant search to kick in.
# this is necessary so super short nicknames like "ed" doesn't flag every three-letter word
# with a "e" in it
redaction_shrinkage = 0.9
# shrink the size of the redaction rectangle so that we redact less of the original text
# and don't accidentally redact text in a different line
max_number_of_nicknames = 5
# maximum number of nicknames to detect before we stop adding more nicknames

# export parameters
export_time_stamp = True
export_page_count = True
export_suspicious_pages_count = True
export_redaction_count = True
export_suspicions = True
export_total_names_looked_for = True

pymupdf.TOOLS.set_small_glyph_heights(on=True) # gets rid of text less aggressively than if set to "false"

#%% create the output folder, if it doesn't exist already
if not os.path.isdir(output_folder):
    os.mkdir(output_folder)
    
# %% create "output" file, which contains confidence data, debug data, etc.

all_files_begin_time = time.time()

# deal with potentially duplicate files
os.chdir(output_folder)
performance_summary_file_name = "performance_summary"
copy_nums = []
found_copies = False
for file in os.listdir():
    if file[0:19] != "performance_summary":
        continue
    found_copies = True
    start_id  = file.find("(") + 1
    start_end = file.find(")")
    copy_num  = int(file[start_id:start_end])
    copy_nums.append(copy_num)
    
if found_copies:
    performance_summary_file_name += " (" + str(max(copy_nums)+1) + ").xlsx" 
else:
    performance_summary_file_name += " (0).xlsx"
    
performances_summary = xlsxwriter.Workbook(performance_summary_file_name)
worksheet = performances_summary.add_worksheet()

worksheet.write(0,0, "AAMC ID")
worksheet.write(0,1, "name")
col = 2; row = 1
if export_time_stamp:             worksheet.write(0,col, "Time Stamp (s)");  col += 1
if export_page_count:             worksheet.write(0,col, "Page Count");      col += 1
if export_suspicious_pages_count: worksheet.write(0,col, "Number of Suspicious Pages"); col += 1
if export_suspicions:             worksheet.write(0,col, "Suspicions"); col += 1
if export_redaction_count:        worksheet.write(0,col, "Redaction Count"); col += 1
if export_total_names_looked_for: worksheet.write(0,col, "Names Looked For");

performances_summary.close()

# display which file it will be writing the performance statistics to
print("Thank you for using PDF_Redaction, written by your handsome BF.")
print("ID-name pairs will written to '"+performance_summary_file_name+\
      "' in the output folder you specified, but PLEASE REFRAIN FROM CHECKING THIS FILE as the code is running. If you and the code tries to access the file at the same time, it will cause problems :)")

#%% read in the English input file
with open(english_ref_file) as file:
    common_english_words_raw = file.read()
common_english_words_raw = common_english_words_raw.split()
common_english_words = {}
for word in common_english_words_raw:
    common_english_words[re.sub(r'[^a-zA-Z0-9]', '', word).lower()] = 1

#%% some helpful function definitions

def search_for_sensitive_words(page_data,sensitive_word):
    ID_top_coordinate = []
    ID_bottom_coordinate = []
    found_line = False
    for i in range(len(page_data['text'])):
        if "".join(page_data['text'][i].split()).lower().find(sensitive_word) != -1:
            ID_top_coordinate.append(page_data['top'][i])
            ID_bottom_coordinate.append(
                page_data['top'][i] + page_data['height'][i])
            found_line = True
    return [ID_top_coordinate, ID_bottom_coordinate, found_line]

def find_with_one_mistake(text, pattern):
    if len(text) - len(pattern) > maximum_length_difference_tolerance:
        return -1    
    if len(pattern) < minimum_name_length_for_spelling_mistake_tolerant_search:
        return text.find(pattern)
    for i in range(len(text) - len(pattern) + 1):
        mismatches = 0
        for j in range(len(pattern)):
            if text[i+j] != pattern[j]:
                mismatches += 1
                if mismatches > spelling_mistake_tolerance:
                    break
        if mismatches <= spelling_mistake_tolerance:
            return i
    return -1

def shrink_rectangle(rect):
    rect = np.array(rect)
    x1 = rect[0]
    x2 = rect[2]
    y1 = rect[1]
    y2 = rect[3]
    center_y = (y2 + y1)/2
    y1_new = (y1 - center_y) * redaction_shrinkage + center_y
    y2_new = (y2 - center_y) * redaction_shrinkage + center_y
    return pymupdf.Rect(x1, y1_new, x2, y2_new)

def clean_list(my_list):
    my_list = [x.lower() for x in my_list] # convert to lower case
    my_list = list(set(my_list))
    while("" in my_list):
        my_list.remove("")
    my_list = [x for x in my_list if len(x) > 1]
    return my_list

# %% obtain relevant text data (names, AAMCID, etc.)

for file_num, file_name in enumerate(os.listdir(input_folder)):
    total_name_of_files = len(os.listdir(input_folder))
    print("----------------------------------------------------------")
    print("Analyzing file ", file_num+1, " of ", total_name_of_files, "--", file_name)
    file_path = input_folder + file_name
    
    begin_time = time.time()
    
    # our debug variable, "suspicions"
    suspicions = ""
    
    # first, flatten all pdfs
    fillpdfs.flatten_pdf(file_path, file_path, as_images=False)
    
    # Convert PDF to images
    pages = convert_from_path(file_path, dpi_used, poppler_path=poppler_path,
                              first_page=0, last_page=1)  # 300 DPI for better quality
    page = pages[0]
    
    # get the text
    page_string = pytesseract.image_to_string(page)
    page_data = pytesseract.image_to_data(
        page, output_type="dict", config="--psm 11")
    page_words = page_string.split(" ")
    page_lines = page_string.split('\n')
    
    names = []
    identifying_info = []
    AAMCID_list = []
    
    # =============================================================================
    # searching for names with image box technique
    ID_top_coordinate, ID_bottom_coordinate, found_name_line = \
        search_for_sensitive_words(page_data,"name")
    if found_name_line == True:
        for i in range(len(ID_top_coordinate)):
            sliced_image = np.array(page)[max(ID_top_coordinate[i]-enlarged_bounds, 0):min(ID_bottom_coordinate[i]+enlarged_bounds, np.shape(np.array(page))[0]),
                                          :]
            name_containing_data = pytesseract.image_to_data(
                sliced_image, output_type='dict')
            add_the_current_word = False
            for j in range(len(name_containing_data['text'])):
                word = name_containing_data['text'][j]
                if add_the_current_word == True:
                    names.append(re.sub(r'[^a-zA-Z0-9]', '', word).lower())
                    names.append(re.sub(r'[^a-zA-Z0-9-]', '', word).lower())
                    if name_containing_data['left'][min(j+1, len(name_containing_data['text'])-1)] - \
                       name_containing_data['left'][j] - name_containing_data['width'][j] > minimum_spacing_needed_before_categories:
                        add_the_current_word = False
                if re.sub(r'[^a-zA-Z0-9]', '', word).lower() == 'name':
                    add_the_current_word = True
                    # but only add the next word if the next word isn't too far away
                    if name_containing_data['left'][min(j+1, len(name_containing_data['text'])-1)] - \
                       name_containing_data['left'][j] - name_containing_data['width'][j] > minimum_spacing_needed_before_categories:
                        add_the_current_word = False
    
    # ============================================================================
    # searching for emails with image box technique
    ID_top_coordinate, ID_bottom_coordinate, found_email_line = \
        search_for_sensitive_words(page_data,"email")
    if found_email_line == True:
        for i in range(len(ID_top_coordinate)):
            sliced_image = np.array(page)[max(ID_top_coordinate[i]-enlarged_bounds, 0):min(ID_bottom_coordinate[i]+enlarged_bounds, np.shape(np.array(page))[0]),
                                          :]
            email_containing_data = pytesseract.image_to_data(
                sliced_image, output_type='dict')
            add_the_current_word = False
            for j in range(len(email_containing_data['text'])):
                word = email_containing_data['text'][j]
                if add_the_current_word == True:
                    identifying_info.append(re.sub(r'[^a-zA-Z0-9]', '', word).lower())
                    identifying_info.append(re.sub(r'[^a-zA-Z0-9-]', '', word).lower())
                    if email_containing_data['left'][min(j+1, len(email_containing_data['text'])-1)] - \
                       email_containing_data['left'][j] - email_containing_data['width'][j] > minimum_spacing_needed_before_categories:
                        add_the_current_word = False
                if re.sub(r'[^a-zA-Z0-9]', '', word).lower() == 'email':
                    add_the_current_word = True
                    # but only add the next word if the next word isn't too far away
                    if email_containing_data['left'][min(j+1, len(email_containing_data['text'])-1)] - \
                       email_containing_data['left'][j] - email_containing_data['width'][j] > minimum_spacing_needed_before_categories:
                        add_the_current_word = False
    
    if len(names) > 0:
        names = list(set(names))
        print("Image recognition identified elements in the name to be", names)
    else:
        print("Image recognition failed to identify the elements in the name. Proceeding with using raw text instead.")
    
    # %%
    # =============================================================================
    # searching for names using raw text. This is often pretty reliable too, lol.
    if len(names) == 0:
        doc = pymupdf.open(file_path)
        page = doc[0]
        page_text_raw = page.get_text()
        page_text_raw_by_lines = page_text_raw.split("\n")
        names = page_text_raw_by_lines[0].split()
    
        email = page_text_raw_by_lines[1]
        names.append(email)
        names.append(email.split('@')[0])
        
    # =============================================================================
    # Break up names with Hyphens, in case that becomes an issue later down the line.
    sub_names = []
    for name in names:
        for name_substring in name.split("-"):
            sub_names.append(name_substring.lower())
    names.extend(sub_names)
    
    if len(names) == 0:
        print("Name identification failed. Using the name of the PDF file instead.")
        
    # =============================================================================
    # searching for names by first using the name of the pdf so we are not moving
    # forward with the analysis with nothing in the names vector
    for word in file_name[0:-4].split():
        names.append(word)
    
    # cleaning up the names list so we have less values to iterate through / search through
    names.extend(identifying_info)
    names = clean_list(names)
        
    # store the original length of "names" vector
    original_number_of_names = len(names)
    
    # printing out what we found the names to be
    print("The following names (or elements of names) were found with raw text.",
          names)
    # %% redact the document, first using the raw PDF text
    doc = pymupdf.open(file_path)
    
    print("Beginning Redaction.")
    redaction_count = 0
    number_of_suspicious_pages = 0
    
    # perform complete redaction
    for page_num, page in enumerate(tqdm(doc)):
        if page_num < page_begin:
            continue
        if page_num == page_end:
            break
    
        redaction_count_per_page = 0
        page.clean_contents()
    
        # prepare the page for adding nicknames / middle names
        nicknames = []
                    
        # -------------------------------------------------------------------------
        # =========================================================================
        # -------------------------------------------------------------------------
        # NAME REDACTION BY SCOURING IMAGE DATA; ALSO, REDACT FACES
        # first, scour the image data for names and faces:
        # first: if the page doesn't contain any useable text, or there is very little
        # text, then we consider the page to be mainly images. In this case, we render
        # the entire page out as one image and conduct our analysis (which might be faster and better)
    
        # if there are any images, we render out the entire page as a single image, and go through the
        # redaction process.
        # regardless of the page, let's use OCR to try to find an AAMC ID. Once
        # we've found it, we can stop this process
        image_list = page.get_images()
        if len(image_list) > 0:
            rendered_page = page.get_pixmap(dpi=dpi_used)
            shape = (rendered_page.height, rendered_page.width, rendered_page.n)
            array_image = np.frombuffer(
                rendered_page.samples, dtype=np.uint8).copy().reshape(shape).squeeze()
    
            # =====================================================================
            # first, redact any facial data detection for images
            # Load pre-trained Haar cascade XML file
            page_rectangle = pymupdf.Rect(page.rect)
            page_rectangle = np.array(page_rectangle)
    
            i1, j1, i2, j2 = page_rectangle
    
            scale_x = (i2-i1)/rendered_page.width
            scale_y = (j2-j1)/rendered_page.height
            trs_x = i1
            trs_y = j1
            
            face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    
            # Detect faces
            faces = face_cascade.detectMultiScale(array_image, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
    
            for i in range(len(faces)):
                x1 = faces[i, 0] * scale_x + trs_x
                x2 = (faces[i, 0] + faces[i, 2]) * scale_x + trs_x
                y1 = faces[i, 1] * scale_y + trs_y
                y2 = (faces[i, 1] + faces[i, 3]) * scale_y + trs_y
                redaction_area = pymupdf.Rect(x1, y1, x2, y2)
                # redaction_area += pymupdf.Rect(pix.x, pix.y, pix.x, pix.y)
                page.add_redact_annot(pymupdf.Rect(x1, y1, x2, y2), fill=[0, 0, 0])
                redaction_count_per_page += 1
    
            # =====================================================================
            # now moving onto detecting text!!
            # correct for image orientation. if this fails, skip the page
            try:
                osd_result = pytesseract.image_to_osd(array_image, output_type='dict')
            except:
                continue
            if osd_result['rotate'] == 180:
                osd_result['rotate'] = 0
    
            # must set the rotation to its current rotation + detected misalignment
            page.set_rotation(page.rotation + osd_result['rotate'])
    
            # render out the page again (the rotated version, if it has been rotated)
            if osd_result['rotate'] != 0:
                rendered_page = page.get_pixmap(dpi=dpi_used)
                shape = (rendered_page.height,
                         rendered_page.width, rendered_page.n)
                array_image = np.frombuffer(
                    rendered_page.samples, dtype=np.uint8).copy().reshape(shape).squeeze()
    
            # convert the image in to B & W
            if len(np.shape(array_image)) > 2:
                array_image = cv2.cvtColor(array_image, cv2.COLOR_BGR2GRAY)
    
            # recompute coordinate transforms
            page_rectangle = pymupdf.Rect(page.rect)
            page_rectangle = np.array(page_rectangle)
    
            i1, j1, i2, j2 = page_rectangle
    
            scale_x = (i2-i1)/rendered_page.width
            scale_y = (j2-j1)/rendered_page.height
            trs_x = i1
            trs_y = j1
    
            # try to extract text data. If this fails, skip the page
            try:
                image_data = pytesseract.image_to_data(
                    array_image, output_type='dict', config="--psm 11")
            except:
                continue
    
            # =================================================================
            # Trying to find the AAMC ID on this page if we haven't found one yet
            # =================================================================
            if len(AAMCID_list) == 0:
                ID_top_coordinate, ID_bottom_coordinate, found_AAMCID_line = \
                    search_for_sensitive_words(image_data,"aamc")
                if found_AAMCID_line == True:
                    for i in range(len(ID_top_coordinate)):
                        sliced_image = array_image[max(ID_top_coordinate[i]-enlarged_bounds, 0):min(ID_bottom_coordinate[i]+enlarged_bounds, np.shape(array_image)[0])]
                        aamc_containing_string = pytesseract.image_to_string(sliced_image)
                        aamc_containing_string = "".join(
                            aamc_containing_string.split()).lower()
                        
                        encountered_numbers = False
                        AAMCID = ''
                        id_start = aamc_containing_string.find("aamcid")
                        for j in range(len(aamc_containing_string[id_start::])):
                            if aamc_containing_string[id_start+j].isdigit():
                                AAMCID = aamc_containing_string[id_start+j:id_start+j+8]
                                break
                        if len(AAMCID) > 0: AAMCID_list.append(AAMCID)
                    if all(i == AAMCID_list[0] for i in AAMCID_list) and len(AAMCID_list) > 0:
                        print("\nImage recognition identified AAMCID to be " + AAMCID_list[0] + " on page " + str(page_num+1) + ".")
                        AAMCID = AAMCID_list[0]
                    else:
                        print("\n2Image recognition failed to identify AAMC ID on page "+str(page_num+1)+".")
        
            # =====================================================================
            # text data detection for images for names
            # scour it for single instances of the name
            for i in range(len(image_data['text'])):
                cleaned_word = re.sub(r'[^a-zA-Z0-9]', '',
                                      image_data['text'][i]).lower()
                for name in names:
                    cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                    if find_with_one_mistake(cleaned_word, cleaned_name) == -1:
                        continue  # if we can't find anything, continue
                    x1 = image_data['left'][i] * scale_x + trs_x
                    x2 = (image_data['left'][i] +
                          image_data['width'][i]) * scale_x + trs_x
                    y1 = image_data['top'][i] * scale_y + trs_y
                    y2 = (image_data['top'][i] +
                          image_data['height'][i]) * scale_y + trs_y
                    redaction_area = pymupdf.Rect(
                        x1, y1, x2, y2) * page.derotation_matrix
                    page.add_redact_annot(shrink_rectangle(redaction_area), fill=[0, 0, 0])
                    redaction_count_per_page += 1
    
                    # detect if there is a word stuck between two "name" words
                    ID_top_coordinate = image_data['top'][i]
                    ID_bottom_coordinate = image_data['top'][i] + image_data['height'][i]
                    words_on_the_same_line = array_image[max(ID_top_coordinate - enlarged_bounds, 0):min(
                        ID_bottom_coordinate + enlarged_bounds, np.shape(array_image)[0]), :]
                    same_line_data = pytesseract.image_to_data(words_on_the_same_line, output_type='dict')
    
                    # now looping through same_line_data and see if we spot any new
                    # "middle names" that can be added to our list
                    for j in range(len(same_line_data['text'])-2):
                        cleaned_word = re.sub(r'[^a-zA-Z0-9]', '', same_line_data['text'][j]).lower()
                        for name in names:
                            cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                            if find_with_one_mistake(cleaned_word, cleaned_name) == -1:
                                continue  # if the first word isn't a name, continue with checking the next name
                            
                            next_word = same_line_data['text'][j+1]                            
                            
                            # if the word is "Mr", "Ms", "Mrs", or "Dr", skip it
                            if re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'ms' or \
                                re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'mr' or \
                                re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'mrs' or \
                                re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'dr':
                                break  # seeing if this word is a nickname if it's one of the above
    
                            # don't add it as a nickname if it's already something we are searching for
                            next_word_cleaned = re.sub(r'[^a-zA-Z0-9]', '', next_word).lower()
                            already_a_name = False
                            for name in names:
                                cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                                if find_with_one_mistake(next_word_cleaned, cleaned_name) != -1:
                                    already_a_name = True; break  # skip it if it's already in the names (so something we are already looking for)
                            if already_a_name: break
                        
                            # don't add it if it's a common word
                            if next_word in common_english_words: break
    
                            # don't add it if we have too many nicknames
                            if len(names) > original_number_of_names + max_number_of_nicknames: break
    
                            two_words_over = same_line_data['text'][j+2]
                            cleaned_two_words_over = re.sub(r'[^a-zA-Z0-9]', '', two_words_over).lower()
                            
                            for name in names:
                                # looping through names again to see if the second word
                                # finds anything
                                cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                                if find_with_one_mistake(cleaned_two_words_over, cleaned_name) == -1:
                                    continue  # go with the next name if the current one isn't it
    
                                # redact the first instsance of the name
                                x1 = same_line_data['left'][j+1] * scale_x + trs_x
                                x2 = (same_line_data['left'][j+1] + same_line_data['width'][j+1]) * scale_x + trs_x
                                y1 = (same_line_data['top'][j+1] + ID_top_coordinate) * scale_y + trs_y
                                y2 = (same_line_data['top'][j+1] + ID_top_coordinate + same_line_data['height'][j+1]) * scale_y + trs_y
                                redaction_area = pymupdf.Rect(x1, y1, x2, y2) * page.derotation_matrix
                                page.add_redact_annot(shrink_rectangle(redaction_area), fill=[0, 0, 0])
                                redaction_count_per_page += 1
    
                                # redact the second instance of the name
                                x1 = same_line_data['left'][j+2] * scale_x + trs_x
                                x2 = (same_line_data['left'][j+2] + same_line_data['width'][j+2]) * scale_x + trs_x
                                y1 = (same_line_data['top'][j+2] + ID_top_coordinate) * scale_y + trs_y
                                y2 = (same_line_data['top'][j+2] + ID_top_coordinate + same_line_data['height'][j+2]) * scale_y + trs_y
                                redaction_area = pymupdf.Rect(x1, y1, x2, y2) * page.derotation_matrix
                                page.add_redact_annot(shrink_rectangle(redaction_area), fill=[0, 0, 0])
                                redaction_count_per_page += 1
    
                                if len(next_word_cleaned) <= 1:
                                    continue  # skip appending it if it's only one letter
                                # append the nicknames to the list of things to redact
                                nicknames = [next_word_cleaned]
                                print("\nImage-based nickname finder found '"+next_word_cleaned+"' to be a nickname. Added to list.")
                                break
                        names.extend(nicknames)
                        names = clean_list(names)
    
        # clear the nicknames list
        nicknames = []
        # -------------------------------------------------------------------------
        # =========================================================================
        # -------------------------------------------------------------------------
        # NAME REDACTION USING RAW TEXT DATA THAT'S IN THE PDF ITSELF
        # First, scour the text data for modified versions of the name, and redact those
        # note: this implementation is mildly inefficient, as it performs a new re-
        # daction for every new instance of the name. But seeing as how each name
        # only appears 5-6 times a page max, it doesn't seem that bad. Computers are
        # quite fast, after all, yo.
        # NOTE: WE HAVE SPECIFICALLY GOTTEN RID OF HYPHENS HERE IN THE CASE THAT
        # A NAME CONTAINS HYPHENS IN THE PAGE THAT'S NOT RECOGNIZED BY OUR NAMES
        # LIST. But we don't append it to names, because this method will always
        # catch something like this
        page_text_raw = page.get_text()
        page_text_raw_by_lines = page_text_raw.split("\n")
        for line in page_text_raw_by_lines:
            for i, word in enumerate(line.split()):
                cleaned_word = re.sub(r'[^a-zA-Z0-9]', '', word).lower()
                for name in names:
                    cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                    if find_with_one_mistake(cleaned_word, cleaned_name) == -1:
                        continue
                    instances = page.search_for(word)
                    for inst in instances:
                        page.add_redact_annot(shrink_rectangle(inst), fill=[0, 0, 0])
                        redaction_count_per_page += 1
                    break;
    
        # =========================================================================
        # first, scour the text data for initials in the name, and get rid of those
        # too. Also, while we are at it, look for any potential nicknames. We define
        # a nickname as a word that's sandwiched by two "name" words and encased in either
        # brackets () [] {} or quotation marks "" ''
        # the searching for initials is PURELY CONTEXTUAL, while, for nicknames,
        # if we find one, we add it to the list. Assert that nicknames cannot be
        # Ms. Dr. Mrs. or Mr.
        page_text_by_words = page_text_raw.split()
        for i, word in enumerate(page_text_by_words):
            cleaned_word = re.sub(r'[^a-zA-Z0-9]', '', word).lower()
            for name in names:
                cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                if find_with_one_mistake(cleaned_word, cleaned_name) == -1:
                    continue  # go to the next name if we can't find anything
                
                # begin initials detection. Start by searching what the next word is
                next_word = page_text_by_words[min(i+1, len(page_text_by_words)-1)]
                
                # if the word is "Mr", "Ms", "Mrs", or "Dr", skip it. Our method
                # looks at the character count so it's easy to mistaken these
                # "words" as being nicknames and/or initials
                if re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'ms' or \
                    re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'mr' or \
                    re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'mrs' or \
                    re.sub(r'[^a-zA-Z0-9]', '', next_word).lower() == 'dr':
                    break  # seeing if this word is a nickname if it's one of the above

                # basically, if the current word is a name, and the next word is
                # only a single letter, then we consider that to be an initial
                # of the person, and we redact it too.
                next_word_contains_1_letters = len(re.sub(r'[^a-zA-Z0-9]', '', next_word)) == 1
                next_word_contains_2_letters = len(re.sub(r'[^a-zA-Z0-9]', '', next_word)) == 2
                last_letter_of_the_next_word_is_a_comma_or_period = any(punctuation in next_word for punctuation in ".,")
                if (next_word_contains_1_letters or next_word_contains_2_letters) \
                        and last_letter_of_the_next_word_is_a_comma_or_period:
                    instances = page.search_for(next_word)
                    for inst in instances:
                        page.add_redact_annot(shrink_rectangle(inst), fill=[0, 0, 0])
                        redaction_count_per_page += 1
                        
                # basically, if the current word is a name, and the word after
                # the word after that is also a name, and the word these two
                # names sandwich is encased in either brackets or quotation marks,
                # we consider this a "nickname" and add it to the redactions list
                two_words_over = page_text_by_words[min(i+2, len(page_text_by_words)-1)]
                cleaned_two_words_over = re.sub(r'[^a-zA-Z0-9]', '', two_words_over).lower()
    
                # uncomment if this takes out too many words because it thinks some random ass word is a nickname
                # contains_brackets = any(punctuation in next_word for punctuation in "()[]{}\"\'“”‘’〈〉《》【】")


                # make sure the current word isn't already a name before we spend
                # effort redacting it
                next_word_cleaned = re.sub(r'[^a-zA-Z0-9]', '', next_word).lower()
                already_a_name = False
                for name in names:
                    cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                    if find_with_one_mistake(next_word_cleaned, cleaned_name) != -1:
                        already_a_name = True; break  # skip it if it's already in the names (so something we are already looking for)
                if already_a_name: break
    
                # don't add it if it's a common word
                if next_word_cleaned in common_english_words: break
    
                # don't add it if we have too many nicknames
                if len(names) > original_number_of_names + max_number_of_nicknames: break
    
                for name in names:
                    cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
                    # if cleaned_two_words_over.find(re.sub(r'[^a-zA-Z0-9]','',name).lower()) != -1 and contains_brackets:
                    if find_with_one_mistake(cleaned_two_words_over, cleaned_name) == -1 or \
                        next_word_contains_1_letters == True or \
                        len(next_word) <= 1:
                        continue  # go to the next name if we don't find  anything
                    nicknames =[next_word_cleaned]
                    print("\nText-based nickname finder found '"+next_word_cleaned+"' to be a nickname. Added to list.")
                    instances = page.search_for(next_word)
                    for inst in instances:
                        page.add_redact_annot(shrink_rectangle(inst), fill=[0, 0, 0])
                        redaction_count_per_page += 1
                    break;
                names = clean_list(names)
                break
             
        #  =========================================================================
        # last, scour the raw text for mentions of the name. This method has the
        # advantage of being able to catch names hyphenated to the next line.
        # we also have the updated names list, so it will catch any nicknames
        # and modified names that it just found, also
        for name in names:
            instances = page.search_for(name)
            # Redact each instance of "Jane Doe" on the current page
            for inst in instances:
                page.add_redact_annot(shrink_rectangle(inst), fill=[0, 0, 0])
                redaction_count_per_page += 1
        
        # =========================================================================
        # search the page for aamc id, and try to append it
        if len(AAMCID_list) == 0:
            AAMCID = ''
            for i, word in enumerate(page_text_by_words):
                cleaned_word = re.sub(r'[^a-zA-Z0-9]', '', word).lower()
                if cleaned_word.find("aamc") == -1:
                    continue  # go to the next name if we can't find anything
                try:
                    potential_aamc_id = re.sub(r'[^a-zA-Z0-9]', '', page_text_by_words[i+1]).lower()
                    if len(potential_aamc_id) == 8 and potential_aamc_id.isdigit():
                        AAMCID = potential_aamc_id
                        print("\nText based AAMC ID Identification found AAMC ID to be: "+AAMCID+" on page "+str(page_num+1)+".")
                        AAMCID_list.append(AAMCID)
                    potential_aamc_id = re.sub(r'[^a-zA-Z0-9]', '', page_text_by_words[i+2]).lower()
                    if len(potential_aamc_id) == 8 and potential_aamc_id.isdigit():
                        AAMCID = potential_aamc_id
                        print("\nText based AAMC ID Identification found AAMC ID to be: "+AAMCID+" on page "+str(page_num+1)+".")
                        AAMCID_list.append(AAMCID)
                    potential_aamc_id = re.sub(r'[^a-zA-Z0-9]', '', page_text_by_words[i+3]).lower()
                    if len(potential_aamc_id) == 8 and potential_aamc_id.isdigit():
                        AAMCID = potential_aamc_id
                        print("\nText based AAMC ID Identification found AAMC ID to be: "+AAMCID+" on page "+str(page_num+1)+".")
                        AAMCID_list.append(AAMCID)
                except:
                    pass
        
        # Apply the redactions to the current page
        try:
            page.apply_redactions()
        except:
            suspicions += "Unable to apply redactions on page " + str(page_num + 1) + ". Human operator presence requested.\n"
            
        # clean up the list
        sub_names = []
        for name in names:
            for name_substring in name.split("-"):
                sub_names.append(name_substring.lower())
        names.extend(sub_names)
        names.extend(nicknames)
        names = clean_list(names)
            
        # calculate final redaction count
        redaction_count += redaction_count_per_page
        
        # determine if this page is suspicious
        if redaction_count_per_page < minimum_number_of_redactions_before_lifting_suspicion:
            number_of_suspicious_pages += 1
    
    # %% redact the document, now using the extracted image text
    
    # first, determine if we have obtained an AAMCID:
    if len(AAMCID_list) == 0:
        AAMCID = "no ID found " + str(random.randint(1000,9999))
        suspicions += "no AAMC ID found! \n"
    
    # create the file name and save it
    document_name = output_folder + AAMCID + ".pdf"
    doc.save(document_name)
    doc.close()
    
    # calculate elapsed time
    end_time = time.time()
    all_files_time_elapsed = end_time
    
    # print redaction statistics
    print("redacted document written to",  AAMCID +
          ".pdf. Time taken for this file:", end_time - begin_time, "s.")

    # calculate suspicions
    if number_of_suspicious_pages / (page_num + 1) > 0.5:
        suspicions += "Too many pages with too few redactions.\n"
    if len(names) == original_number_of_names + max_number_of_nicknames:
        suspicions += "Too many nicknames added.\n"

    # write the results to output file
    wb = load_workbook(performance_summary_file_name)
    new_data = []

    new_data.append(AAMCID)
    new_data.append(file_name[0:-4])
    if export_time_stamp:             
        new_data.append(round(all_files_time_elapsed-all_files_begin_time,2))
    if export_page_count:             
        new_data.append(page_num + 1)
    if export_suspicious_pages_count:             
        new_data.append(number_of_suspicious_pages)
    if export_suspicions:
        new_data.append(suspicions)
    if export_redaction_count:        
        new_data.append(redaction_count)
    if export_total_names_looked_for: 
        new_data.append(str(names));
    new_data = [new_data]
        
    ws = wb.active
    for row in new_data:
        ws.append(row)

    wb.save(performance_summary_file_name)
    wb.close()

# %% print final summary statistics
print("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
print("Finished processing all files. Total of", total_name_of_files, 
  "files have been processed, average time of",
  round((all_files_time_elapsed-all_files_begin_time)/total_name_of_files,2),"seconds per file.",
  "\nMore performance statistics can be found in '"+performance_summary_file_name+"'.")
print("=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=")
print("Hey bb! Thanks for using my code! MWah <3 MWah <3 MWah <3")
